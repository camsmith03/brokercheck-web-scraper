import pandas as pd
import numpy as np
import time
import tkinter as tk
import random
from tkinter import filedialog
from tkinter.filedialog import asksaveasfilename
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Color, fills
from openpyxl.utils.dataframe import dataframe_to_rows


def getLinks(fileName, columnName):
    """Generates the links by parsing the csv file and getting
       the CRD numbers to append to the link headers.

    Args:
        fileName (String): location of CSV file
        columnName (String): column containing the CRD number

    Returns:
        list: list of links to scrape
    """
    df = pd.read_csv(fileName)
    crds = df[columnName]
    
    header = "https://brokercheck.finra.org/individual/summary/"
    
    links = []
    
    for crd in crds:
        if not np.isnan(crd):
            links.append(header + str(crd))
    
    return links


def scrapeData(links):
    """Uses selenium webdriver to scrape each site one by one,
       gathering the data into responses and returning the result.

    Args:
        links (list): list of links to scrape

    Returns:
        list: raw html data from the scraped links
    """
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    driver = webdriver.Chrome(options=options)

    responses = []
    
    approximateTime = 11.3 + ((len(links) - 1) * 2.82)
    
    if approximateTime < 60:
        print("\n\nScraping data (Approx time: {} seconds)...\n"
              .format(str(round(approximateTime))))
    elif approximateTime < 3600:
        print("\n\nScraping data (Approx time: {} minute(s) {} seconds)...\n"
              .format(str(round(approximateTime // 60)), str(round(approximateTime % 60))))
    else:
        tempSeconds = approximateTime % (24 * 3600)
        print("\n\nScraping data (Approx time: {} hour(s) {} minute(s) {} seconds)...\n"
              .format(str(round(tempSeconds // 3600)), str(round((tempSeconds % 3600) // 60)), 
                      str(round((tempSeconds % 3600) % 60))))


    for link in links:
        timeout = False
        print("Scraping CRD: " + link.rsplit('/', 1)[1])
        driver.get(link)

        try:
            element = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/bc-root/div/bc-individual-container-page/bc-individual-detail-page/div[2]/investor-tools-individual-summary-template/div/div[1]/div[1]/div/investor-tools-big-name/div[1]/span[1]"))
            )
        except TimeoutException:
            timeout = True
        finally:
            if not timeout:
                responses.append(driver.page_source)
            else:
                responses.append((int(link.rsplit('/', 1)[1]),"IA"))
                timeout = False
        
        time.sleep(random.randint(1, 3))

    driver.quit()
    
    print("\nScraping complete, processing data...\n")

    return responses
  

def parseData(responses):
    """Takes in the raw set of data and parses it into a list of
       dictionaries, each dictionary containing the data for one
       broker.

    Args:
        responses (list): raw scraped data

    Returns:
        list: clean set of data to be exported
    """
    print("\nParsing data...\n")
    scraped_data_raw = []

    for response in responses:
        scraped_data_partial = {}
        if type(response) == tuple:
            scraped_data_partial["Broker CRD"] = response[0]
            scraped_data_partial["Is a Broker"] = "null"
            scraped_data_partial["Is an Investment Adviser"] = True
        else:
            soup = BeautifulSoup(response, features="html.parser")
            scraped_data_partial["name"] = soup.find("span", {"class": "text-lg sm:text-sm font-semibold"})
            scraped_data_partial["status"] = soup.find_all("span", {"class": "text-gray-80 text-xs font-medium"})
            scraped_data_partial["crd"] = soup.find("div", {"class": "text-gray-85 text-left font-semibold mt-2 text-sm ng-star-inserted"})
            scraped_data_partial["firm"] = soup.find("div", {"class": "flex flex-col text-sm"})
            scraped_data_partial["background"] = soup.find_all("div", {"class": "flex-1 flex flex-col justify-center"})
        
        scraped_data_raw.append(scraped_data_partial)
        
        
    scraped_data_clean = []

    for scrape in scraped_data_raw:
        if len(scrape) > 3:
            clean_data = {}
            clean_data["Broker Name"] = scrape["name"].string.strip()

            raw_status = scrape["status"]
            
            if len(raw_status) == 1:
                if scrape["status"][0].find("span").string == "Broker":
                    clean_data["Is a Broker"] = True
                else:
                    clean_data["Is a Broker"] = False
                
                clean_data["Is an Investment Adviser"] = False
            else:
                if raw_status[1].find("span").string == "Broker":
                    clean_data["Is a Broker"] = True
                else:
                    clean_data["Is a Broker"] = False
                
                if raw_status[0].find("span", {"title": "Investment Adviser"}).string.strip() == "Investment Adviser":
                    clean_data["Is an Investment Adviser"] = True
                else:
                    clean_data["Is an Investment Adviser"] = False
            
            clean_data["Broker CRD"] = int(scrape["crd"].find("span").next_sibling.string) 
            
            if scrape['firm']:
                clean_data["Firm Name"] = scrape['firm'].find("span").string
                clean_data["Firm CRD"] = int(scrape['firm'].find("span").next_sibling.find("span").next_sibling.string)

                rawAddress = scrape['firm'].find("investor-tools-address")

                rawStreetAddress = rawAddress.next_element

                for x in range(3):
                    rawStreetAddress = rawStreetAddress.next_element

                clean_data["Firm Street"] = rawStreetAddress.strip()

                rawCityStateZip = rawAddress.find("br")

                for x in range(4):
                    rawCityStateZip = rawCityStateZip.next_element

                rawCityStateZip = rawCityStateZip.strip()

                rawStateZip = rawCityStateZip.split(" ", 1)[1].split(" ", 1)

                clean_data["Firm State"] = rawStateZip[0]

                clean_data["Firm Zip"] = rawStateZip[1]
            else:
                clean_data["Firm Name"] = "none"
                clean_data["Firm CRD"] = "none"
                clean_data["Firm Street"] = "none"
                clean_data["Firm State"] = "none"
                clean_data["Firm Zip"] = "none"
                
            clean_data["Number of Disclosures"] = int(scrape["background"][0].find("span", {"class": "sm:text-lg sm:font-semibold text-3xl ng-star-inserted"}).string.strip())
            
            rawYearsFirms = scrape["background"][1].find_all("span", {"class": "sm:text-lg sm:font-semibold text-3xl ng-star-inserted"})

            if len(rawYearsFirms) == 2:
                clean_data["Years of Experience"] = int(rawYearsFirms[0].string.strip())
                clean_data["Number of Firms"] = rawYearsFirms[1].string.strip()
            else:
                clean_data["Years of Experience"] = int(rawYearsFirms[0].string.strip())
                clean_data["Number of Firms"] = scrape["background"][1].find("span", {"class": "sm:text-lg sm:font-semibold text-xl ng-star-inserted"}).string.strip()
            
            scraped_data_clean.append(clean_data)
        else:
            scraped_data_clean.append(scrape)
            
    return scraped_data_clean


def writeData(scraped_data_clean, directory):
    """Takes in the cleaned data and converts it into an excel
       file with highlighting on whether each one is a registered
       broker.

    Args:
        scraped_data_clean (list): list of clean data
        directory (string): location of where to save file
    """
    print("\nWriting data to Excel file...\n")
    completed_df = pd.DataFrame(scraped_data_clean)
    falseIndexes = list(completed_df.loc[(completed_df["Is a Broker"] == False) 
                                        & (completed_df["Is an Investment Adviser"] == False)].index.values)
    onlyIAs = list(completed_df.loc[(completed_df["Is a Broker"] == False) 
                                        & (completed_df["Is an Investment Adviser"] == True)].index.values)
    neverBrokers = list(completed_df.loc[(completed_df["Is a Broker"] == "null")].index.values)
    

    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(completed_df, index=False, header=True):
        ws.append(r)

    for index in falseIndexes:
        for cell in ws[str(index + 2) + ":" + str(index + 2)]:
            cell.fill = fills.PatternFill(patternType='solid', fgColor=Color("FF0000"))

    for index in onlyIAs:
        for cell in ws[str(index + 2) + ":" + str(index + 2)]:
            cell.fill = fills.PatternFill(patternType='solid', fgColor=Color("FFFF00"))

    for index in neverBrokers:
        for cell in ws[str(index + 2) + ":" + str(index + 2)]:
            cell.fill = fills.PatternFill(patternType='solid', fgColor=Color("A020F0"))
            
    wb.save(directory)
    
    print("\nDone! Check the directory for the file named {}\n".format(directory.rsplit("/", 1)[-1]))
    


def main():
    """Runs all of the necessary functions and passes them all of
       the important data.    
    """
    print("BrokerCheck Web Scraper\nBy Cameron Smith\n================================")
    
    root = tk.Tk()
    root.withdraw()
    filetypes = (("Text CSV", "*.csv"), ("All files", "*.*"))
    file_path = filedialog.askopenfilename(filetypes=filetypes)
    root.destroy()
    columnName = input("Enter the column name that contains the CRD numbers: ")
    links = getLinks(file_path, columnName)
    responses = scrapeData(links)
    scraped_data_clean = parseData(responses)
    
    print("\nSelect a location to save the file.\n")
    newRoot = tk.Tk()
    newRoot.withdraw()
    save_location = asksaveasfilename(initialfile = "Untitled.xlsx", defaultextension=" .xlsx", filetypes=[("Excel 2007-365", "*.xlsx"), ("All files", "*.txt")])
    newRoot.destroy()
    
    writeData(scraped_data_clean, save_location)
    
    
    
    userExit = input("\nPress enter to exit...")


if __name__ == '__main__':
    main()